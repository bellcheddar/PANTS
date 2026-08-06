"""Ingest the Science 2025 PET depolymerase landscape library.

*Landscape profiling of PET depolymerases using a natural sequence cluster framework*,
Science 2025, doi:10.1126/science.adp5637, supplementary Data S3.

2,064 library entries, of which 183 were expressed and assayed for product release. The
183 are what is ingested: an entry with no activity value was not measured, and a protein
nobody measured is not evidence of anything, in either direction.

THE THRESHOLD COMES FROM THE DATA, NOT FROM ME. Every assayed entry carries its own
replicate standard deviation, so "did this enzyme do anything" can be asked against the
assay's own noise instead of against a number chosen to make the counts look good:

    active            activity > 2 x its own STDEV      102 entries
    measured inactive activity <= its own STDEV          69 entries
    ambiguous         between the two                    12 entries

The ambiguous band is ingested with its measurements and labelled NEITHER. Twelve enzymes
sitting between one and two standard deviations of zero are exactly the ones a threshold
would be tuned on, and calling them either way would assert more than the experiment
supports. They are marked excluded from training with the reason recorded, so they are
counted and visible rather than quietly dropped.

The 64 entries at or below zero activity are the reason this file matters. PAZy records
only what worked, so its negatives mean "not reported active"; these were expressed,
assayed and found not to release product.
"""

from __future__ import annotations

import hashlib
import pathlib
import re
from typing import Any, Dict, List, Optional, Tuple

from ..db import connect, now
from ..db.manifest import stage_manifest

STAGE = "science_landscape"
DOI = "10.1126/science.adp5637"
CITATION = ("Landscape profiling of PET depolymerases using a natural sequence cluster "
            "framework, Science 2025 (doi:10.1126/science.adp5637), Data S3")

# Product release was measured on PET film under one protocol; the paper reports it in
# micromolar of released product, so the group is one assay and pooling within it is safe.
GROUP = "science2025|PET|product_release_uM"

REPLICATE_COLS = [f"Product Release_{i}  (μM)" for i in range(1, 12)]
# Column 1 is spelled with a single space where the rest use two. Written out rather than
# generated, because a generated name that silently misses turns a replicate into a None.
REPLICATE_COLS[0] = "Product Release_1 (μM)"


def _sid(sequence: str) -> str:
    return "SCI:" + hashlib.sha1(sequence.encode()).hexdigest()[:10]


def _num(v: Any) -> Optional[float]:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load(path: pathlib.Path) -> List[dict]:
    """Assayed entries only, each with its sequence, activity, noise and melting points."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    ix = {h: i for i, h in enumerate(rows[0])}

    out: List[dict] = []
    for r in rows[1:]:
        activity = _num(r[ix["Activity"]])
        seq = r[ix["AAseq"]]
        if activity is None or not isinstance(seq, str) or len(seq) < 50:
            continue
        tms = [_num(r[ix[k]]) for k in ("TmB", "TmD") if k in ix]
        tms = [t for t in tms if t is not None]
        name = r[ix["Reported Name"]] if "Reported Name" in ix else None
        name = name.strip() if isinstance(name, str) and name.strip() not in ("", "-") else None
        out.append({
            "library_id": r[ix["Library ID"]],
            "accession": r[ix["Accession code"]],
            "phylum": r[ix["*Major phylum"]],
            "sequence": re.sub(r"[^A-Z]", "", seq.upper()),
            "activity_uM": activity,
            "stdev": _num(r[ix["Activity_STDEV"]]),
            "tm_c": round(sum(tms) / len(tms), 2) if tms else None,
            "reported_name": name,
            "replicates": [_num(r[ix[c]]) for c in REPLICATE_COLS if c in ix],
        })
    return out


def classify(entry: dict) -> str:
    """active | inactive | ambiguous, against the entry's own replicate noise."""
    a, s = entry["activity_uM"], entry["stdev"]
    if s is None:
        return "active" if a > 0 else "inactive"
    if a > 2 * s:
        return "active"
    if a <= s:
        return "inactive"
    return "ambiguous"


def ingest(path: pathlib.Path, label: str = "v1") -> Dict[str, Any]:
    entries = load(path)
    report: Dict[str, Any] = {"assayed": len(entries), "active": 0, "inactive": 0,
                              "ambiguous": 0, "new_rows": 0, "matched_existing": 0,
                              "measurements": 0, "named": 0}

    with stage_manifest(STAGE, label=label) as m:
        with connect() as c:
            existing = {row[0]: row[1] for row in c.execute(
                "SELECT sequence, enzyme_id FROM characterised_enzymes "
                "WHERE sequence IS NOT NULL")}

            for e in entries:
                verdict = classify(e)
                report[verdict] += 1
                seq = e["sequence"]
                enzyme_id = existing.get(seq)
                if enzyme_id:
                    report["matched_existing"] += 1
                else:
                    enzyme_id = _sid(seq)
                    report["new_rows"] += 1
                if e["reported_name"]:
                    report["named"] += 1

                sd = f" +/- {e['stdev']:.3g}" if e["stdev"] is not None else ""
                note = (
                    f"Library entry {e['library_id']} ({e['accession']}) in the Science 2025 "
                    f"PET depolymerase landscape. Product release "
                    f"{e['activity_uM']:.3g}{sd} uM, classed {verdict.upper()} against its "
                    f"own replicate noise (active above 2 SD, inactive at or below 1 SD). "
                    + {"inactive": "This is a MEASURED negative: expressed, assayed, and no "
                                   "product released above noise. ",
                       "ambiguous": "Between one and two standard deviations of zero, so "
                                    "labelled neither way and excluded from training. ",
                       "active": ""}[verdict]
                    + f"Source: {CITATION}.")

                c.execute(
                    "INSERT INTO characterised_enzymes "
                    "(enzyme_id, uniprot, common_name, sequence, seq_length, family, "
                    " taxonomy_lineage, is_positive, is_negative, is_near_miss, "
                    " within_family_basis, excluded_from_training, exclusion_reason, "
                    " activity_substrate_notes, source_ref, primary_doi, added_at) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
                    "ON CONFLICT(enzyme_id) DO UPDATE SET "
                    " activity_substrate_notes=excluded.activity_substrate_notes, "
                    " common_name=COALESCE(characterised_enzymes.common_name, excluded.common_name), "
                    " primary_doi=COALESCE(characterised_enzymes.primary_doi, excluded.primary_doi), "
                    " within_family_basis=COALESCE(characterised_enzymes.within_family_basis, "
                    "                              excluded.within_family_basis)",
                    (enzyme_id, None, e["reported_name"], seq, len(seq), "petase_like",
                     e["phylum"],
                     1 if verdict == "active" else 0,
                     0,
                     1 if verdict == "inactive" else 0,
                     "measured-inactive" if verdict == "inactive" else None,
                     1 if verdict == "ambiguous" else 0,
                     "activity within 2 SD of zero" if verdict == "ambiguous" else None,
                     note, "Science-landscape-measured", DOI, now()))

                c.execute(
                    "INSERT INTO activity_measurements "
                    "(enzyme_id, substrate_form, parameter_type, rate_value, rate_units, "
                    " product_measured, raw_text, evidence_code, comparable_group_id, "
                    " source_doi, extracted_at, extraction_confidence) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (enzyme_id, "PET film", "product_release", e["activity_uM"], "uM",
                     "TPA + MHET",
                     f"product release {e['activity_uM']:.4g}{sd} uM "
                     f"({len([x for x in e['replicates'] if x is not None])} replicates)",
                     "ECO:0000269", GROUP, DOI, now(), "high"))
                report["measurements"] += 1

                if e["tm_c"]:
                    c.execute(
                        "INSERT INTO activity_measurements "
                        "(enzyme_id, parameter_type, rate_value, rate_units, raw_text, "
                        " evidence_code, comparable_group_id, source_doi, extracted_at, "
                        " extraction_confidence) VALUES (?,?,?,?,?,?,?,?,?,?)",
                        (enzyme_id, "tm", e["tm_c"], "degC",
                         f"melting temperature {e['tm_c']} degC",
                         "ECO:0000269", "science2025|tm", DOI, now(), "high"))
                    report["measurements"] += 1
            c.commit()

        m.counts(n_input=len(entries),
                 n_output=report["new_rows"] + report["matched_existing"],
                 n_discarded=report["ambiguous"])
    return report
