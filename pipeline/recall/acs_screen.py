"""Ingest the ACS Catalysis 2025 high-throughput PET hydrolase screen.

Norton-Baker, Komp, Gado et al., *Machine Learning-Guided Identification of PET Hydrolases
from Natural Diversity*, ACS Catal. 2025, 15, 16070-16083, doi:10.1021/acscatal.5c03460.
Source data deposited openly with the paper (Supporting Information `cs5c03460_si_002.xlsx`).

This matters more than its size. Every evaluation in this project has been limited by one
thing: the negative class. PAZy records only positive substrate associations, so "not
reported active on PET" cannot distinguish an enzyme that was tested and failed from one
nobody tested, and the catalogue holds 29 within-family negatives on that weak basis.

This screen expressed and assayed a panel under ONE protocol and reports what did not
work. Proteins measured below the detection floor here are inactive because somebody
measured them, which is a different and much stronger claim, and it is the claim the
evaluation needs.

Two design decisions worth stating, because both could have been made carelessly:

  A protein is only labelled from a condition it was actually assayed under. The screen
  ran at 40 and 60 degC, two substrates and five pH values, and a protein absent from a
  condition is not a zero in it. `comparable_group_id` therefore keys on substrate,
  temperature and pH together, the same discipline the existing measurements use, so a
  crystalline-powder result at 60 degC is never pooled with an amorphous-film result at 40.

  Activity is taken from percent depolymerisation, not from a normalised rate. The
  normalised columns divide by enzyme loading, which is right for comparing catalytic
  efficiency and wrong for asking "did this enzyme do anything", since a tiny loading
  turns noise into a large per-mg number. The threshold is stated, applied once, and
  recorded on every row so it can be moved.
"""

from __future__ import annotations

import hashlib
import json
import pathlib
import re
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

from ..db import connect, now
from ..db.manifest import stage_manifest

STAGE = "acs_screen"
DOI = "10.1021/acscatal.5c03460"
CITATION = ("Norton-Baker, Komp, Gado et al. 2025, ACS Catal. 15:16070-16083 "
            "(doi:10.1021/acscatal.5c03460)")

# Below this the assay did not distinguish the protein from background. Stated rather than
# tuned: it is the paper's own reporting floor for a positive hit, and every measurement
# row carries the value used so a later reader can move it and recount.
ACTIVE_PCT = 0.1

SUBSTRATE_LABEL = {"cryPow": "crystalline PET powder", "aFilm": "amorphous PET film"}

# The screen's accession column holds whatever database the sequence came from: UniProt for
# some, NCBI WP_/MBQ_/KAA_ for most, MGnify MGYP for a few, and the literal string "NA".
# Only a real UniProt accession may go in the `uniprot` column, because everything
# downstream treats that column as one -- AlphaFold fetches by it, the deposit linker
# queries it, the fragment check sends it to UniProt's accession filter. Writing an NCBI
# id there made UniProt reject the whole batched query with a 400.
UNIPROT_ACCESSION = re.compile(
    r"^([OPQ][0-9][A-Z0-9]{3}[0-9]|[A-NR-Z][0-9]([A-Z][A-Z0-9]{2}[0-9]){1,2})(-\d+)?$")


def _sid(sequence: str) -> str:
    """Stable id from the sequence itself, so a re-run cannot mint duplicate rows."""
    return "ACS:" + hashlib.sha1(sequence.encode()).hexdigest()[:10]


def load_workbook(path: pathlib.Path) -> Tuple[Dict[str, dict], List[dict]]:
    """(proteins by sample_id, activity rows). Requires openpyxl, which is offline-only."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    rows = list(wb["Rounds"].iter_rows(values_only=True))
    ix = {h: i for i, h in enumerate(rows[0])}
    proteins: Dict[str, dict] = {}
    for r in rows[1:]:
        sid = r[ix["sample_id"]]
        seq = r[ix["seq_aa"]]
        if not sid or not isinstance(seq, str) or len(seq) < 50:
            continue
        seq = re.sub(r"[^A-Z]", "", seq.upper())
        tms = [r[ix[k]] for k in ("tm_1", "tm_2")]
        tms = [float(t) for t in tms if isinstance(t, (int, float))]
        acc = r[ix["accession"]]
        acc = acc.strip() if isinstance(acc, str) and acc.strip() not in ("", "None", "NA") else None
        proteins.setdefault(sid, {
            "sample_id": sid, "sequence": seq,
            "accession": acc,
            "uniprot": acc if acc and UNIPROT_ACCESSION.match(acc) else None,
            "tm_c": round(sum(tms) / len(tms), 1) if tms else None,
            "round": r[ix["round"]],
        })

    arows = list(wb["Activity"].iter_rows(values_only=True))
    ai = {h: i for i, h in enumerate(arows[0])}
    keep = ("sample_id", "substrate", "temperature_c", "pH", "percent_depolymerization",
            "umol_product_per_mg_enzyme", "enzyme_loading_ug", "reaction_vol_mL", "dataset")
    activity = [{k: r[ai[k]] for k in keep if k in ai} for r in arows[1:]]
    return proteins, [a for a in activity if a.get("sample_id")]


def summarise(activity: List[dict]) -> Dict[str, dict]:
    """Best result per (protein, condition), and the protein's best across all conditions."""
    per_cond: Dict[Tuple, dict] = {}
    for a in activity:
        pct = a.get("percent_depolymerization")
        if not isinstance(pct, (int, float)):
            continue
        key = (a["sample_id"], a.get("substrate"), a.get("temperature_c"), a.get("pH"))
        cur = per_cond.get(key)
        if cur is None or pct > cur["pct"]:
            per_cond[key] = {"pct": float(pct),
                             "rate": a.get("umol_product_per_mg_enzyme"),
                             "loading_ug": a.get("enzyme_loading_ug")}
    by_protein: Dict[str, dict] = defaultdict(lambda: {"best": None, "conditions": []})
    for (sid, sub, temp, ph), v in per_cond.items():
        rec = by_protein[sid]
        rec["conditions"].append({"substrate": sub, "temperature_c": temp, "ph": ph, **v})
        if rec["best"] is None or v["pct"] > rec["best"]:
            rec["best"] = v["pct"]
    return dict(by_protein)


def ingest(path: pathlib.Path, label: str = "v1", active_pct: float = ACTIVE_PCT
           ) -> Dict[str, Any]:
    proteins, activity = load_workbook(path)
    summary = summarise(activity)

    report: Dict[str, Any] = {
        "proteins_in_file": len(proteins),
        "assayed": len(summary),
        "active": 0, "inactive": 0, "not_assayed": 0,
        "new_rows": 0, "matched_existing": 0, "measurements": 0,
        "active_at_40c": 0, "threshold_pct": active_pct,
    }

    with stage_manifest(STAGE, label=label) as m:
        with connect() as c:
            existing = {row[0]: row[1] for row in c.execute(
                "SELECT sequence, enzyme_id FROM characterised_enzymes "
                "WHERE sequence IS NOT NULL")}

            for sid, p in sorted(proteins.items()):
                seq = p["sequence"]
                s = summary.get(sid)
                if s is None:
                    report["not_assayed"] += 1
                    continue
                is_active = s["best"] >= active_pct
                report["active" if is_active else "inactive"] += 1
                if any(cd["temperature_c"] == 40 and cd["pct"] >= active_pct
                       for cd in s["conditions"]):
                    report["active_at_40c"] += 1

                enzyme_id = existing.get(seq)
                if enzyme_id:
                    report["matched_existing"] += 1
                else:
                    enzyme_id = _sid(seq)
                    report["new_rows"] += 1

                note = (
                    f"{p['sample_id']} ({p['accession'] or 'no accession'}) in the ACS "
                    f"Catalysis 2025 high-throughput screen. "
                    f"Best result {s['best']:.3f}% depolymerisation across "
                    f"{len(s['conditions'])} assayed conditions, "
                    f"{'ACTIVE' if is_active else 'MEASURED INACTIVE'} at a "
                    f"{active_pct}% threshold. "
                    + ("This is a measured negative: the protein was expressed and assayed "
                       "and did not depolymerise PET, which is a stronger statement than "
                       "PAZy's 'not reported active'. " if not is_active else "")
                    + f"Source: {CITATION}.")

                c.execute(
                    "INSERT INTO characterised_enzymes "
                    "(enzyme_id, uniprot, sequence, seq_length, family, "
                    " is_positive, is_negative, is_near_miss, within_family_basis, "
                    " activity_substrate_notes, source_ref, primary_doi, added_at) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?) "
                    "ON CONFLICT(enzyme_id) DO UPDATE SET "
                    " activity_substrate_notes=excluded.activity_substrate_notes, "
                    " primary_doi=COALESCE(characterised_enzymes.primary_doi, excluded.primary_doi), "
                    " within_family_basis=COALESCE(characterised_enzymes.within_family_basis, "
                    "                              excluded.within_family_basis)",
                    (enzyme_id, p["uniprot"], seq, len(seq), "petase_like",
                     1 if is_active else 0,
                     # NOT is_negative: that tier means a different fold entirely. These are
                     # polyesterase-family proteins expressed and assayed and found not to
                     # depolymerise PET, which is exactly the near-miss definition the head
                     # is meant to learn -- the boundary INSIDE the family, not outside it.
                     0,
                     0 if is_active else 1,
                     # And the strongest basis this column has carried: measured, not
                     # inferred from a database that only records what worked.
                     None if is_active else "measured-inactive",
                     note, "ACS-screen-measured", DOI, now()))

                for cd in s["conditions"]:
                    group = (f"acs2025|{cd['substrate']}|{cd['temperature_c']}C|pH{cd['ph']}")
                    c.execute(
                        "INSERT INTO activity_measurements "
                        "(enzyme_id, substrate_form, temperature_c, ph, parameter_type, "
                        " rate_value, rate_units, product_measured, raw_text, evidence_code, "
                        " comparable_group_id, source_doi, extracted_at, extraction_confidence) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (enzyme_id, SUBSTRATE_LABEL.get(cd["substrate"], cd["substrate"]),
                         cd["temperature_c"], cd["ph"], "percent_depolymerization",
                         cd["pct"], "%", "TPA + MHET",
                         f"{cd['pct']:.4f}% depolymerisation of "
                         f"{SUBSTRATE_LABEL.get(cd['substrate'], cd['substrate'])} at "
                         f"{cd['temperature_c']} degC, pH {cd['ph']}",
                         "ECO:0000269", group, DOI, now(), "high"))
                    report["measurements"] += 1

                if p["tm_c"]:
                    c.execute(
                        "INSERT INTO activity_measurements "
                        "(enzyme_id, parameter_type, rate_value, rate_units, raw_text, "
                        " evidence_code, comparable_group_id, source_doi, extracted_at, "
                        " extraction_confidence) VALUES (?,?,?,?,?,?,?,?,?,?)",
                        (enzyme_id, "tm", p["tm_c"], "degC",
                         f"melting temperature {p['tm_c']} degC (nanoDSF)",
                         "ECO:0000269", "acs2025|tm", DOI, now(), "high"))
                    report["measurements"] += 1
            c.commit()

        m.counts(n_input=len(proteins), n_output=report["new_rows"] + report["matched_existing"],
                 n_discarded=report["not_assayed"])
    return report
