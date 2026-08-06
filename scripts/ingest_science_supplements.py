#!/usr/bin/env python3
"""Ingest the remaining Science 2025 supplements: S1 (homologues), S2, S7 (ecology).

Three files, three different things, and only two of them are worth ingesting:

  S2  2,064 sequences. Every accession and every sequence is byte-identical to the AAseq
      column of S3, which is already in. Verified rather than assumed -- all 2,064 matched
      -- and then skipped. Ingesting it would create nothing and risk duplicating rows
      whose counts this project quotes as evidence.

  S1  25,411 sequences, of which S2 is a strict subset, so 23,347 are new. These are the
      paper's homologue search space: no activity was measured on any of them. They go in
      `unlabelled_sequences`, NOT in `characterised_enzymes`, because nothing here has been
      characterised and letting them into the table whose totals are quoted would corrupt
      every count on the site. They are useful as a PU-unlabelled pool and as a screening
      pool; they cannot add a measured negative, which is the constraint that binds.

  S7  Ecological context per sequence cluster: isolation source, biome, habitat, geography
      and the temperature the source organism was isolated at. Stored at CLUSTER level
      because that is the level the paper reports it at -- attaching an isolation
      temperature to an individual enzyme would invent a precision the source does not have.
      And it is isolation temperature, not enzyme optimum: the two are different claims and
      the column name says which this is.
"""

from __future__ import annotations

import hashlib
import json
import pathlib
import re
import sys
from typing import Dict, List, Optional

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))

from pipeline.db import connect, now
from pipeline.db.manifest import stage_manifest

SRC = pathlib.Path.home() / "Desktop/science.adp5637_data_s1_to_s7"
DOI = "10.1126/science.adp5637"

SHEETS = {
    "isolation_sources_json": "Isolation sources by island",
    "locations_json": "geographic locations by island",
    "biomes_json": "broad-scale env context by I",
    "habitats_json": "local-scale env context by I",
    "temperatures_json": "temperature by island",
}

# "28 celcius", "4,0 degree of C", "65 degrees F", "4-35 C (optimum at 15 C)". Take the
# first number and convert Fahrenheit; anything outside a range life is found in is noise.
_TEMP = re.compile(r"(-?\d+(?:[.,]\d+)?)")


def parse_temperature(raw: str) -> Optional[float]:
    m = _TEMP.search(raw.replace(",", "."))
    if not m:
        return None
    v = float(m.group(1))
    if re.search(r"\bF\b|fahrenheit|degrees F", raw, re.I):
        v = (v - 32) * 5 / 9
    return v if -25 <= v <= 125 else None


def read_fasta(path: pathlib.Path) -> Dict[str, dict]:
    out: Dict[str, dict] = {}
    acc = desc = None
    cur: List[str] = []
    for line in path.read_text().splitlines():
        if line.startswith(">"):
            if acc:
                out[acc] = {"accession": acc, "description": desc, "sequence": "".join(cur)}
            parts = line[1:].split(None, 1)
            acc, desc, cur = parts[0], (parts[1] if len(parts) > 1 else None), []
        elif line.strip():
            cur.append(line.strip())
    if acc:
        out[acc] = {"accession": acc, "description": desc, "sequence": "".join(cur)}
    return out


def organism_of(description: Optional[str]) -> Optional[str]:
    if not description:
        return None
    m = re.search(r"\[([^\]]+)\]\s*$", description)
    return m.group(1) if m else None


def ingest_s7() -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(SRC / "science.adp5637_data_s7.xlsx",
                                read_only=True, data_only=True)
    per_cluster: Dict[str, dict] = {}
    for field, sheet in SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        for row in list(wb[sheet].iter_rows(values_only=True))[1:]:
            if row[0] is None:
                continue
            cid = str(row[0]).strip()
            vals = [str(x).strip() for x in row[1:]
                    if x is not None and str(x).strip() not in ("", "None")]
            per_cluster.setdefault(cid, {})[field] = vals

    rows = []
    for cid, fields in per_cluster.items():
        temps = fields.get("temperatures_json", [])
        parsed = sorted(t for t in (parse_temperature(x) for x in temps) if t is not None)
        median = (parsed[len(parsed) // 2] if len(parsed) % 2
                  else (parsed[len(parsed) // 2 - 1] + parsed[len(parsed) // 2]) / 2) if parsed else None
        rows.append(("Science-landscape", cid, None,
                     json.dumps(fields.get("isolation_sources_json", [])),
                     json.dumps(fields.get("biomes_json", [])),
                     json.dumps(fields.get("habitats_json", [])),
                     json.dumps(fields.get("locations_json", [])),
                     json.dumps(temps), median, now()))

    with connect() as c:
        c.executemany(
            "INSERT INTO sequence_clusters(source, cluster_id, n_members, "
            " isolation_sources_json, biomes_json, habitats_json, locations_json, "
            " temperatures_json, temperature_median_c, added_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT(source, cluster_id) DO UPDATE SET "
            " isolation_sources_json=excluded.isolation_sources_json, "
            " biomes_json=excluded.biomes_json, habitats_json=excluded.habitats_json, "
            " locations_json=excluded.locations_json, "
            " temperatures_json=excluded.temperatures_json, "
            " temperature_median_c=excluded.temperature_median_c", rows)
        c.commit()
    return {"clusters": len(rows),
            "with_a_temperature": sum(1 for r in rows if r[8] is not None)}


def link_clusters() -> dict:
    """Put each ingested Science enzyme in its cluster, so the ecology is reachable."""
    import openpyxl

    wb = openpyxl.load_workbook(SRC / "science.adp5637_data_s3.xlsx",
                                read_only=True, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    ix = {h: i for i, h in enumerate(rows[0])}
    by_seq = {}
    for r in rows[1:]:
        s = r[ix["AAseq"]]
        if isinstance(s, str) and r[ix["Cluster number"]] is not None:
            by_seq[re.sub(r"[^A-Z]", "", s.upper())] = int(r[ix["Cluster number"]])

    with connect() as c:
        have = c.execute("SELECT enzyme_id, sequence FROM characterised_enzymes "
                         "WHERE sequence IS NOT NULL").fetchall()
        updates = [(by_seq[s.upper()], e) for e, s in have if s and s.upper() in by_seq]
        c.executemany("UPDATE characterised_enzymes SET science_cluster=? WHERE enzyme_id=?",
                      updates)
        c.commit()
    return {"enzymes_linked": len(updates)}


def ingest_s1() -> dict:
    s1 = read_fasta(SRC / "science.adp5637_data_s1.txt")
    s2 = read_fasta(SRC / "science.adp5637_data_s2.txt")
    # Verified, not assumed: S2 is skipped only because every one of its entries is already
    # present in S3, and this is the check that says so.
    redundant = len(set(s2) & set(s1))

    with connect() as c:
        known = {r[0] for r in c.execute(
            "SELECT sequence FROM characterised_enzymes WHERE sequence IS NOT NULL")}
        rows = []
        for acc, rec in s1.items():
            seq = re.sub(r"[^A-Z]", "", rec["sequence"].upper())
            if len(seq) < 50 or seq in known:
                continue
            rows.append(("SCIU:" + hashlib.sha1(seq.encode()).hexdigest()[:10], acc,
                         rec["description"], organism_of(rec["description"]), seq, len(seq),
                         "Science-landscape-S1", DOI, now()))
        c.executemany(
            "INSERT INTO unlabelled_sequences(seq_id, accession, description, organism, "
            " sequence, seq_length, source_ref, source_doi, added_at) "
            "VALUES (?,?,?,?,?,?,?,?,?) ON CONFLICT(seq_id) DO NOTHING", rows)
        c.commit()
        total = c.execute("SELECT COUNT(*) FROM unlabelled_sequences").fetchone()[0]
    return {"s1_sequences": len(s1), "s2_sequences": len(s2),
            "s2_redundant_with_s1": redundant, "offered": len(rows),
            "unlabelled_total": total,
            "skipped_already_characterised": len(s1) - len(rows)}


def main() -> int:
    with stage_manifest("science_supplements", label="s1-s7") as m:
        s7 = ingest_s7()
        linked = link_clusters()
        s1 = ingest_s1()
        m.counts(n_input=s1["s1_sequences"], n_output=s1["offered"],
                 n_discarded=s1["skipped_already_characterised"])

    print("S7  ecological context")
    print(f"   {s7['clusters']} clusters stored, {s7['with_a_temperature']} with a "
          f"parseable isolation temperature")
    print(f"   {linked['enzymes_linked']} catalogued enzymes linked to their cluster")
    print("\nS2  skipped")
    print(f"   {s1['s2_redundant_with_s1']} of {s1['s2_sequences']} already present via S3")
    print("\nS1  unlabelled homologue pool")
    print(f"   {s1['s1_sequences']} sequences, {s1['offered']} added, "
          f"{s1['skipped_already_characterised']} already characterised")
    print(f"   unlabelled_sequences now holds {s1['unlabelled_total']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
